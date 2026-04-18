import csv
import os

try:
    import openpyxl
except ImportError:  # optional dependency at runtime
    openpyxl = None

import logging
logger = logging.getLogger("SKILLAB-Projector")


class EscoLoader:
    def __init__(self, engine):
        self.engine = engine

    def _load_skill_uris_from_csv(self, path: str) -> set[str]:
        uris = set()
        if not os.path.exists(path):
            logger.warning(f"CSV non trovato: {path}")
            return uris

        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                uri = (row.get("conceptUri") or "").strip()
                if uri:
                    uris.add(uri)

        logger.info(f"Caricati {len(uris)} URI da {path}")
        return uris

    def _normalize_nace_lookup_code(self, raw_code: str) -> str:
        raw = str(raw_code or "").strip()
        if not raw:
            return ""

        if "/" in raw:
            raw = raw.rstrip("/").split("/")[-1]

        raw = raw.upper().replace(" ", "")
        if len(raw) == 1 and raw.isalpha():
            return raw

        cleaned = "".join(ch for ch in raw if ch.isdigit() or ch == ".")
        if not cleaned:
            return ""

        if "." in cleaned:
            head, tail = cleaned.split(".", 1)
            head = head[:2]
            tail = "".join(ch for ch in tail if ch.isdigit())
            if not head:
                return ""
            if not tail:
                return head
            if len(tail) == 1:
                return f"{head}.{tail}"
            return f"{head}.{tail[:2]}"

        digits = "".join(ch for ch in cleaned if ch.isdigit())
        if not digits:
            return ""

        if len(digits) <= 2:
            return digits.zfill(2)
        if len(digits) == 3:
            return f"{digits[:2]}.{digits[2]}"
        return f"{digits[:2]}.{digits[2:4]}"

    def load_nace_labels(self, filename: str = "nace_codes_2_1.csv"):
        path = os.path.join(os.getcwd(), "complementary_data", filename)
        if not os.path.exists(path):
            logger.warning(f"NACE codes file not found: {path}")
            return

        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    label = str(row.get("Activity") or "").strip()
                    if not label:
                        continue

                    for level_key, csv_key in (
                        ("section", "Section"),
                        ("division", "Division"),
                        ("group", "Group"),
                        ("class", "Class"),
                    ):
                        code = self._normalize_nace_lookup_code(row.get(csv_key))
                        if not code:
                            continue
                        self.engine.nace_labels[code] = label
                        self.engine.nace_labels_by_level[level_key][code] = label

            logger.info(f"Loaded NACE labels: {len(self.engine.nace_labels)}")
        except Exception as e:
            logger.warning(f"Could not load NACE labels CSV: {e}")

    def _register_occupation_nace_mapping(self, occ_uri: str, nace_code: str, nace_label: str):
        occ_uri = str(occ_uri or "").strip()
        nace_code = self._normalize_nace_lookup_code(nace_code)
        nace_label = str(nace_label or "").strip()
        if not occ_uri or not nace_code:
            return

        if nace_label:
            self.engine.nace_labels[nace_code] = nace_label

        entries = self.engine.occupation_nace_map[occ_uri]
        if any(e.get("code") == nace_code for e in entries):
            return

        entries.append({
            "code": nace_code,
            "label": self.engine.nace_labels.get(nace_code, nace_label or nace_code)
        })

    def load_esco_nace_crosswalk(self, filename: str = "ESCO-NACE rev. 2.1 crosswalk (1).xlsx"):
        """
        Preferred source for ESCO occupation -> NACE mappings and NACE labels.
        Falls back silently if workbook is missing.
        """
        path = os.path.join(os.getcwd(), "complementary_data", filename)
        if not os.path.exists(path):
            # try common alternate filename
            alt_path = os.path.join(os.getcwd(), "complementary_data", "ESCO-NACE rev. 2.1 crosswalk.xlsx")
            if os.path.exists(alt_path):
                path = alt_path
            else:
                logger.warning(f"ESCO-NACE crosswalk workbook not found: {path}")
                return

        if openpyxl is None:
            logger.warning("openpyxl is not installed: ESCO-NACE crosswalk loading is disabled")
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(c or "").strip() for c in next(rows)]

            def idx(name: str) -> int:
                for i, h in enumerate(header):
                    if h.lower() == name.lower():
                        return i
                return -1

            i_esco_uri = idx("ESCO URI")
            i_nace_code = idx("NACE code")
            i_nace_title = idx("NACE title")

            if i_esco_uri < 0 or i_nace_code < 0:
                logger.warning("Crosswalk workbook missing expected columns (ESCO URI / NACE code)")
                return

            for row in rows:
                if not row:
                    continue
                occ_uri = row[i_esco_uri] if i_esco_uri < len(row) else ""
                nace_code = row[i_nace_code] if i_nace_code < len(row) else ""
                nace_title = row[i_nace_title] if i_nace_title >= 0 and i_nace_title < len(row) else ""
                self._register_occupation_nace_mapping(occ_uri, nace_code, nace_title)

            logger.info(
                f"Loaded ESCO-NACE crosswalk: {len(self.engine.occupation_nace_map)} occupations, "
                f"{len(self.engine.nace_labels)} NACE labels"
            )
        except Exception as e:
            logger.warning(f"Could not load ESCO-NACE crosswalk workbook: {e}")



    def load_local_esco_support(self):

        """
        Load local CSV support files for:
        - occupation metadata
        - skill hierarchy
        - occupation-skill canonical relations

        This step is safe and incremental: if a file is missing, it is skipped.
        """
        self.load_esco_nace_crosswalk()
        self.load_nace_labels()
        base_dir = os.getcwd()

        occupations_file = os.path.join(base_dir, "complementary_data", "occupations_en.csv")
        skills_hierarchy_file = os.path.join(base_dir, "complementary_data", "skillsHierarchy_en.csv")
        occ_skill_rel_file = os.path.join(base_dir, "complementary_data", "occupationSkillRelations_en.csv")
        isco_groups_file = os.path.join(base_dir, "complementary_data", "ISCOGroups_en.csv")
        skill_groups_file = os.path.join(base_dir, "complementary_data", "skillGroups_en.csv")

        # 1) Occupation metadata
        if os.path.exists(occupations_file):
            try:
                with open(occupations_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        occ_id = (
                            row.get("conceptUri")
                            or row.get("id")
                            or row.get("occupationUri")
                            or ""
                        ).strip()
                        if not occ_id:
                            continue

                        self.engine.occupation_meta[occ_id] = {
                            "label": (row.get("preferredLabel") or row.get("label") or "").strip(),
                            "isco_group": (row.get("iscoGroup") or "").strip(),
                            "nace_code": (row.get("naceCode") or "").strip(),
                            "raw": row,
                        }

                        # Fallback mapping when crosswalk workbook is not available.
                        self._register_occupation_nace_mapping(
                            occ_uri=occ_id,
                            nace_code=row.get("naceCode"),
                            nace_label=self.engine.nace_labels.get(
                                self._normalize_nace_lookup_code(row.get("naceCode")),
                                ""
                            )
                        )
                logger.info(f"Loaded occupation metadata: {len(self.engine.occupation_meta)}")
            except Exception as e:
                logger.warning(f"Could not load occupations_lt.csv: {e}")

        # 2) Skill hierarchy
        if os.path.exists(skills_hierarchy_file):
            try:
                with open(skills_hierarchy_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        skill_id = (
                                row.get("conceptUri")
                                or row.get("id")
                                or row.get("skillUri")
                                or ""
                        ).strip()
                        if not skill_id:
                            continue

                        level_1 = (row.get("level1") or row.get("ESCO Level 1") or row.get("skillLevel1") or "").strip()
                        level_2 = (row.get("level2") or row.get("ESCO Level 2") or row.get("skillLevel2") or "").strip()
                        level_3 = (row.get("level3") or row.get("ESCO Level 3") or row.get("skillLevel3") or "").strip()

                        self.engine.skill_hierarchy[skill_id] = {
                            "level_1": level_1,
                            "level_2": level_2,
                            "level_3": level_3,
                            "raw": row,
                        }

                logger.info(f"Loaded skill hierarchy: {len(self.engine.skill_hierarchy)}")
            except Exception as e:
                logger.warning(f"Could not load skillsHierarchy_en.csv: {e}")
        # 3) Skill group labels
        if os.path.exists(skill_groups_file):
            try:
                with open(skill_groups_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        gid = (
                            row.get("conceptUri")
                            or row.get("id")
                            or row.get("skillGroup")
                            or ""
                        ).strip()

                        label = (
                            row.get("preferredLabel")
                            or row.get("label")
                            or row.get("title")
                            or ""
                        ).strip()

                        if gid and label:
                            self.engine.skill_group_labels[gid] = label

                            # also store short-code fallback, e.g. .../skill/S4.8 -> S4.8
                            short_gid = gid.rstrip("/").split("/")[-1]
                            if short_gid:
                                self.engine.skill_group_labels[short_gid] = label

                logger.info(f"Loaded skill group labels: {len(self.engine.skill_group_labels)}")
            except Exception as e:
                logger.warning(f"Could not load skillGroups_en.csv: {e}")

        # 4) Occupation-skill canonical relations
        if os.path.exists(occ_skill_rel_file):
            try:
                with open(occ_skill_rel_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        occ_id = (
                            row.get("occupationUri")
                            or row.get("occupation")
                            or row.get("occupation_id")
                            or row.get("conceptUriOccupation")
                            or ""
                        ).strip()

                        skill_id = (
                            row.get("skillUri")
                            or row.get("skill")
                            or row.get("skill_id")
                            or row.get("conceptUriSkill")
                            or ""
                        ).strip()

                        if occ_id and skill_id:
                            # save for each occupations the skill requested
                            self.engine.occ_skill_relations[occ_id].add(skill_id)
                logger.info(f"Loaded canonical occupation-skill relations: {len(self.engine.occ_skill_relations)} occupations")
            except Exception as e:
                logger.warning(f"Could not load occupationSkillRelations_lt.csv: {e}")

        # 4) Optional ISCO group labels
        if os.path.exists(isco_groups_file):
            try:
                with open(isco_groups_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        raw_gid = (
                                row.get("conceptUri")
                                or row.get("id")
                                or row.get("iscoGroup")
                                or ""
                        ).strip()

                        gid = raw_gid.split("/")[-1].replace("C", "")

                        label = (row.get("preferredLabel") or row.get("label") or "").strip()
                        if gid:
                            self.engine.occupation_group_labels[gid] = label
                logger.info(f"Loaded ISCO group labels: {len(self.engine.occupation_group_labels)}")
            except Exception as e:
                logger.warning(f"Could not load ISCOGroups_lt.csv: {e}")

    def load_official_esco_matrix(self, filename: str = "Skills_Occupations Matrix Tables_ESCOv1.2.0_1.xlsx"):
        """
        Load the official ESCO matrix workbook in a lightweight indexed form.

        We extract:
        - available sheets and their meaning from the Overview sheet
        - for each matrix sheet:
            occupation_group_id -> {
                "occupation_group_label": ...,
                "profile": {skill_group_id: share}
            }

        This keeps the implementation incremental and lookup-oriented.
        """
        path = os.path.join(os.getcwd(), "complementary_data", filename)
        if openpyxl is None:
            logger.warning("openpyxl is not installed: official ESCO matrix loading is disabled")
            return

        if not os.path.exists(path):
            logger.warning(f"Official ESCO matrix file not found: {path}")
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # 1. Parse Overview
            if "Overview" in wb.sheetnames:
                ws = wb["Overview"]
                rows = list(ws.iter_rows(values_only=True))
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    sheet_name = str(row[0]).strip()
                    self.engine.esco_matrix_overview[sheet_name] = {
                        "skill_level": row[1],
                        "occupation_level": row[2],
                        "size": row[3],
                    }

            # 2. Parse every matrix sheet
            for sheet_name in wb.sheetnames:
                if sheet_name == "Overview":
                    continue
                if not sheet_name.startswith("Matrix"):
                    continue

                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)

                try:
                    header_1 = next(rows)
                    header_2 = next(rows)
                except StopIteration:
                    continue

                if not header_1:
                    continue

                # skill group ids start from column index 2
                skill_group_ids = []
                for col_idx in range(2, len(header_1)):
                    skill_group_ids.append(
                        str(header_1[col_idx]).strip() if header_1[col_idx] else ""
                    )

                for row in rows:
                    if not row or not row[0]:
                        continue

                    occupation_group_id = str(row[0]).strip()
                    occupation_group_label = str(row[1]).strip() if len(row) > 1 and row[1] else occupation_group_id

                    profile = {}
                    for idx, skill_group_id in enumerate(skill_group_ids, start=2):
                        if not skill_group_id:
                            continue
                        value = row[idx] if idx < len(row) else None
                        if value is None:
                            continue
                        try:
                            profile[skill_group_id] = float(value)
                        except Exception:
                            continue

                    self.engine.esco_matrix_profiles[(sheet_name, occupation_group_id)] = {
                        "occupation_group_label": occupation_group_label,
                        "profile": profile
                    }

            self.engine.esco_matrix_loaded = True
            logger.info(
                f"Loaded official ESCO matrix: "
                f"{len(self.engine.esco_matrix_overview)} sheets in overview, "
                f"{len(self.engine.esco_matrix_profiles)} occupation profiles"
            )

        except Exception as e:
            logger.warning(f"Could not load official ESCO matrix: {e}")
