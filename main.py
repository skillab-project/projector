import csv
import logging
import asyncio
import httpx
import os
import hashlib
import json
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from fastapi import FastAPI, Form
from typing import List, Optional
from dotenv import load_dotenv
import openpyxl
from schemas import ProjectorResponse, EmergingSkillsResponse, StopResponse

# Configurazione Logger
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger("SKILLAB-Projector")

load_dotenv()
app = FastAPI(title="SKILLAB Projector Microservice")


class ProjectorEngine:
    def __init__(self):
        self.api_url = os.getenv("TRACKER_API")
        self.username = os.getenv("TRACKER_USERNAME")
        self.password = os.getenv("TRACKER_PASSWORD")
        self.token = None
        self.skill_map = {}
        self.sector_map = {}  # Cache specifica per le occupazioni/settori
        # --- Local ESCO support maps ---
        self.occupation_meta = {}          # occ_id -> meta info from local CSV
        self.skill_hierarchy = {}          # skill_id -> hierarchy/group info
        self.occ_skill_relations = defaultdict(set)   # occ_id -> set(skill_id)
        self.occupation_group_labels = {}  # optional: group id -> readable label
        self.matrix_profiles = {}          # optional, for later step
        self.occ_skill_observed = defaultdict(Counter)   # occ_id -> Counter(skill_id -> count)
        self.sector_skill_observed = defaultdict(Counter)  # sector -> Counter(skill_id -> count)
        self.sector_skill_canonical = defaultdict(Counter)  # sector -> Counter(skill_id -> count)
        self.sector_skillgroup_observed = defaultdict(Counter)   # sector -> Counter(skill_group -> count)
        self.sector_skillgroup_canonical = defaultdict(Counter)  # sector -> Counter(skill_group -> count)
        self.esco_matrix_overview = {}   # sheet_name -> metadata
        self.esco_matrix_profiles = {}   # (sheet_name, occupation_group_id) -> profile dict
        self.esco_matrix_loaded = False
        self.stop_requested = False
        # Timeout impostato a None per evitare ReadTimeout su query pesanti
        self.client = httpx.AsyncClient(timeout=None)
        self.skill_group_labels = {}   # group_id -> readable label
        #
        # self.green_skill_uris = self._load_skill_uris_from_csv("complementary_data/greenSkillsCollection_lt.csv")
        # self.digital_skill_uris = self._load_skill_uris_from_csv("complementary_data/digitalSkillsCollection_lt.csv")
    def build_canonical_sector_skillgroup_matrix(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 2,
        reset: bool = True
    ):
        if reset:
            self.sector_skillgroup_canonical = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                sector_name = self.get_sector_from_occupation(occ_id, level=sector_level)
                canonical_skills = self.occ_skill_relations.get(occ_id, set())

                for skill_id in canonical_skills:
                    skill_group = self.get_skill_group(skill_id, level=skill_group_level)
                    self.sector_skillgroup_canonical[sector_name][skill_group] += 1


        return self.sector_skillgroup_canonical
    def load_official_esco_matrix(self, filename: str = "Skills_Occupations Matrix Tables_ESCOv1.2.0_1.xlsx"):
        """
        Load the official ESCO matrix workbook in a lightweight indexed form.

        We extract:
        - available sheets and their meaning from the Overview sheet
        - for each matrix sheet:
            occupation_group_id -> {
                "occupation_group_label": ...,
                "profile": {skill_group_id: share}
            }

        This keeps the implementation incremental and lookup-oriented.
        """
        path = os.path.join(os.getcwd(), "complementary_data", filename)
        if not os.path.exists(path):
            logger.warning(f"Official ESCO matrix file not found: {path}")
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # 1. Parse Overview
            if "Overview" in wb.sheetnames:
                ws = wb["Overview"]
                rows = list(ws.iter_rows(values_only=True))
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    sheet_name = str(row[0]).strip()
                    self.esco_matrix_overview[sheet_name] = {
                        "skill_level": row[1],
                        "occupation_level": row[2],
                        "size": row[3],
                    }

            # 2. Parse every matrix sheet
            for sheet_name in wb.sheetnames:
                if sheet_name == "Overview":
                    continue
                if not sheet_name.startswith("Matrix"):
                    continue

                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)

                try:
                    header_1 = next(rows)
                    header_2 = next(rows)
                except StopIteration:
                    continue

                if not header_1:
                    continue

                # skill group ids start from column index 2
                skill_group_ids = []
                for col_idx in range(2, len(header_1)):
                    skill_group_ids.append(
                        str(header_1[col_idx]).strip() if header_1[col_idx] else ""
                    )

                for row in rows:
                    if not row or not row[0]:
                        continue

                    occupation_group_id = str(row[0]).strip()
                    occupation_group_label = str(row[1]).strip() if len(row) > 1 and row[1] else occupation_group_id

                    profile = {}
                    for idx, skill_group_id in enumerate(skill_group_ids, start=2):
                        if not skill_group_id:
                            continue
                        value = row[idx] if idx < len(row) else None
                        if value is None:
                            continue
                        try:
                            profile[skill_group_id] = float(value)
                        except Exception:
                            continue

                    self.esco_matrix_profiles[(sheet_name, occupation_group_id)] = {
                        "occupation_group_label": occupation_group_label,
                        "profile": profile
                    }

            self.esco_matrix_loaded = True
            logger.info(
                f"Loaded official ESCO matrix: "
                f"{len(self.esco_matrix_overview)} sheets in overview, "
                f"{len(self.esco_matrix_profiles)} occupation profiles"
            )

        except Exception as e:
            logger.warning(f"Could not load official ESCO matrix: {e}")

    def get_skill_group(self, skill_id: str, level: int = 2) -> str:
        """
        Resolve the ESCO skill group for a skill using the local skill hierarchy.

        Parameters
        ----------
        skill_id : str
            ESCO skill URI or id
        level : int
            Skill hierarchy level to use: 1, 2, or 3

        Returns
        -------
        str
            Skill group identifier or fallback label
        """
        skill_id = str(skill_id).strip()
        if not skill_id:
            return "Skill group not specified"

        meta = self.skill_hierarchy.get(skill_id)
        if meta:
            if level == 1:
                group_id = (meta.get("level_1") or "").strip()
            elif level == 3:
                group_id = (meta.get("level_3") or "").strip()
            else:
                group_id = (meta.get("level_2") or "").strip()

            if group_id:
                return group_id

        # fallback to resolved skill label if available
        skill_meta = self.skill_map.get(skill_id, {})
        label = skill_meta.get("label", "").strip()
        if label:
            return label

        return "Skill group not specified"
    def build_observed_sector_skillgroup_matrix(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 2,
        reset: bool = True
    ):
        """
        Build a sector -> observed skill-group count matrix from raw jobs.
        """
        if reset:
            self.sector_skillgroup_observed = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                sector_name = self.get_sector_from_occupation(occ_id, level=sector_level)
                for skill_id in job.get("skills", []):
                    skill_id = str(skill_id).strip()
                    if not skill_id:
                        continue

                    skill_group = self.get_skill_group(skill_id, level=skill_group_level)
                    self.sector_skillgroup_observed[sector_name][skill_group] += 1

        return self.sector_skillgroup_observed

    def _read_group_counter(self, counter: Counter, top_k: int = 20):
        """
        Internal helper: convert a Counter into a sorted list with frequencies.
        """
        total = sum(counter.values())
        results = []

        for group_id, count in counter.most_common(top_k):
            results.append({
                "group_id": group_id,
                "group_label": self.get_skill_group_label(group_id),
                "count": count,
                "frequency": round(count / total, 6) if total > 0 else 0.0
            })

        return {
            "total_mentions": total,
            "unique_groups": len(counter),
            "top_groups": results
        }
    def get_skill_group_label(self, group_id: str) -> str:
        """
        Resolve a human-readable label for a skill group id.
        Supports both full ESCO URI and short code.
        """
        group_id = str(group_id).strip()
        if not group_id:
            return "Skill group not specified"

        if group_id in self.skill_group_labels:
            return self.skill_group_labels[group_id]

        if group_id.startswith("http"):
            short_gid = group_id.rstrip("/").split("/")[-1]
            if short_gid in self.skill_group_labels:
                return self.skill_group_labels[short_gid]

        return group_id
    def summarize_observed_sector_skillgroups(self, top_k: int = 20):
        """
        Build a readable summary of observed ESCO skill groups per sector.
        """
        results = []

        for sector_name, counter in self.sector_skillgroup_observed.items():
            base = self._read_group_counter(counter, top_k=top_k)
            results.append({
                "sector": sector_name,
                "total_group_mentions": base["total_mentions"],
                "unique_groups": base["unique_groups"],
                "top_groups": base["top_groups"]
            })

        return sorted(results, key=lambda x: x["total_group_mentions"], reverse=True)
    def summarize_canonical_sector_skillgroups(self, top_k: int = 20):
        """
        Build a readable summary of canonical ESCO skill groups per sector.
        """
        results = []

        for sector_name, counter in self.sector_skillgroup_canonical.items():
            base = self._read_group_counter(counter, top_k=top_k)
            results.append({
                "sector": sector_name,
                "total_group_mentions": base["total_mentions"],
                "unique_groups": base["unique_groups"],
                "top_groups": base["top_groups"]
            })

        return sorted(results, key=lambda x: x["total_group_mentions"], reverse=True)
    def compare_observed_and_canonical_groups_for_sector(
        self,
        sector_name: str,
        top_k: int = 20
    ):
        """
        Compare observed vs canonical skill-group distributions for one sector.
        """
        sector_name = str(sector_name).strip()

        observed_counter = self.sector_skillgroup_observed.get(sector_name, Counter())
        canonical_counter = self.sector_skillgroup_canonical.get(sector_name, Counter())

        observed = self._read_group_counter(observed_counter, top_k=top_k)
        canonical = self._read_group_counter(canonical_counter, top_k=top_k)

        return {
            "sector": sector_name,
            "observed_groups": {
                "total_group_mentions": observed["total_mentions"],
                "unique_groups": observed["unique_groups"],
                "top_groups": observed["top_groups"]
            },
            "canonical_groups": {
                "total_group_mentions": canonical["total_mentions"],
                "unique_groups": canonical["unique_groups"],
                "top_groups": canonical["top_groups"]
            }
        }
    def build_and_summarize_observed_sector_skillgroups(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 2,
        top_k: int = 20,
        reset: bool = True
    ):
        self.build_observed_sector_skillgroup_matrix(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            reset=reset
        )
        return self.summarize_observed_sector_skillgroups(top_k=top_k)


    def build_and_summarize_canonical_sector_skillgroups(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 2,
        top_k: int = 20,
        reset: bool = True
    ):
        self.build_canonical_sector_skillgroup_matrix(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            reset=reset
        )
        return self.summarize_canonical_sector_skillgroups(top_k=top_k)

    def build_observed_occupation_skill_matrix(self, jobs: List[dict], reset: bool = True):
        """
        Build an observed occupation -> skill count matrix from raw Tracker jobs.

        Parameters
        ----------
        jobs : List[dict]
            Raw jobs returned by the Tracker
        reset : bool
            If True, clears the previous matrix before rebuilding it

        Returns
        -------
        defaultdict(Counter)
            occ_id -> Counter(skill_id -> count)
        """
        if reset:
            self.occ_skill_observed = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                for skill_id in job.get("skills", []):
                    skill_id = str(skill_id).strip()
                    if not skill_id:
                        continue
                    self.occ_skill_observed[occ_id][skill_id] += 1

        return self.occ_skill_observed

    def get_canonical_skills_for_occupation(self, occ_id: str, resolve_labels: bool = False, top_k: int = 50):
        """
        Return canonical ESCO skills linked to an occupation through the local
        occupation-skill relations CSV.

        Since the CSV relations are unweighted at this stage, every canonical skill
        gets count=1.
        """
        occ_id = str(occ_id).strip()
        skill_ids = sorted(self.occ_skill_relations.get(occ_id, set()))
        total_skill_mentions = len(skill_ids)

        results = []
        for skill_id in skill_ids[:top_k]:
            entry = {
                "skill_id": skill_id,
                "count": 1,
                "frequency": round(1 / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0,
            }

            if resolve_labels:
                meta = self.skill_map.get(skill_id, {})
                entry["label"] = meta.get("label", skill_id)
                entry["is_green"] = meta.get("is_green", False)
                entry["is_digital"] = meta.get("is_digital", False)

            results.append(entry)

        return results

    def get_sector_label(self, sector_code: str) -> str:
        """
        Resolve a human-readable label for a sector/group code.
        """
        sector_code = str(sector_code).strip()
        if not sector_code:
            return "Sector not specified"

        # direct lookup
        if sector_code in self.occupation_group_labels:
            return self.occupation_group_labels[sector_code]

        # URI form
        uri_key = f"http://data.europa.eu/esco/isco/{sector_code}"
        if uri_key in self.occupation_group_labels:
            return self.occupation_group_labels[uri_key]

        return sector_code

    def build_canonical_sector_skill_matrix(self, jobs: List[dict], sector_level: str = "isco_group", reset: bool = True):
        """
        Build a sector -> canonical skill matrix.

        For each job:
        - extract the primary occupation
        - resolve the sector from that occupation
        - take all canonical ESCO skills linked to that occupation
        - accumulate them at sector level

        Each canonical skill contributes +1 per job occurrence of that occupation.
        """
        if reset:
            self.sector_skill_canonical = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                sector_name = self.get_sector_from_occupation(occ_id, level=sector_level)
                canonical_skills = self.occ_skill_relations.get(occ_id, set())

                for skill_id in canonical_skills:
                    skill_id = str(skill_id).strip()
                    if not skill_id:
                        continue
                    self.sector_skill_canonical[sector_name][skill_id] += 1

        return self.sector_skill_canonical

    def get_canonical_skills_for_sector(self, sector_name: str, resolve_labels: bool = False, top_k: int = 20):
        """
        Return canonical skills for a sector, sorted by count.
        """
        sector_name = str(sector_name).strip()
        counter = self.sector_skill_canonical.get(sector_name, Counter())
        total_skill_mentions = sum(counter.values())

        results = []
        for skill_id, count in counter.most_common(top_k):
            entry = {
                "skill_id": skill_id,
                "count": count,
                "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0,
            }
            if resolve_labels:
                meta = self.skill_map.get(skill_id, {})
                entry["label"] = meta.get("label", skill_id)
                entry["is_green"] = meta.get("is_green", False)
                entry["is_digital"] = meta.get("is_digital", False)
            results.append(entry)

        return results

    def summarize_canonical_sector_skills(
        self,
        resolve_labels: bool = False,
        top_k: int = 20
    ):
        """
        Build a readable summary of canonical ESCO skills per sector, starting from
        self.sector_skill_canonical.
        """
        results = []

        for sector_name, counter in self.sector_skill_canonical.items():
            total_skill_mentions = sum(counter.values())

            top_skills = []
            for skill_id, count in counter.most_common(top_k):
                entry = {
                    "skill_id": skill_id,
                    "count": count,
                    "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0
                }

                if resolve_labels:
                    meta = self.skill_map.get(skill_id, {})
                    entry["label"] = meta.get("label", skill_id)
                    entry["is_green"] = meta.get("is_green", False)
                    entry["is_digital"] = meta.get("is_digital", False)

                top_skills.append(entry)

            results.append({
                "sector": sector_name,
                "total_skill_mentions": total_skill_mentions,
                "unique_skills": len(counter),
                "top_skills": top_skills
            })

        return sorted(results, key=lambda x: x["total_skill_mentions"], reverse=True)
    def compare_observed_and_canonical_for_sector(
        self,
        sector_name: str,
        resolve_labels: bool = False,
        top_k: int = 20
    ):
        """
        Return both observed and canonical summaries for a sector.
        """
        return {
            "sector": sector_name,
            "observed": self.summarize_single_sector(
                sector_name=sector_name,
                resolve_labels=resolve_labels,
                top_k=top_k
            ),
            "canonical": {
                "sector": sector_name,
                "top_skills": self.get_canonical_skills_for_sector(
                    sector_name=sector_name,
                    resolve_labels=resolve_labels,
                    top_k=top_k
                ),
                "total_skill_mentions": sum(self.sector_skill_canonical.get(sector_name, Counter()).values()),
                "unique_skills": len(self.sector_skill_canonical.get(sector_name, Counter()))
            }
        }

    def get_observed_skills_for_sector(self, sector_name: str, resolve_labels: bool = False, top_k: int = 20):
        """
        Return the observed skills for a sector, sorted by count.
        """
        sector_name = str(sector_name).strip()
        counter = self.sector_skill_observed.get(sector_name, Counter())
        total_skill_mentions = sum(counter.values())

        results = []
        for skill_id, count in counter.most_common(top_k):
            entry = {
                "skill_id": skill_id,
                "count": count,
                "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0,
            }
            if resolve_labels:
                meta = self.skill_map.get(skill_id, {})
                entry["label"] = meta.get("label", skill_id)
                entry["is_green"] = meta.get("is_green", False)
                entry["is_digital"] = meta.get("is_digital", False)
            results.append(entry)

        return results

    def build_observed_sector_skill_matrix(self, jobs: List[dict], sector_level: str = "isco_group", reset: bool = True):
        """
        Aggregate observed skills at sector level starting from raw jobs.

        Parameters
        ----------
        jobs : List[dict]
            Raw Tracker jobs
        sector_level : str
            Passed to get_sector_from_occupation()
        reset : bool
            If True, clears the previous sector matrix before rebuilding it

        Returns
        -------
        defaultdict(Counter)
            sector_name -> Counter(skill_id -> count)
        """
        if reset:
            self.sector_skill_observed = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                sector_name = self.get_sector_from_occupation(occ_id, level=sector_level)
                for skill_id in job.get("skills", []):
                    skill_id = str(skill_id).strip()
                    if not skill_id:
                        continue
                    self.sector_skill_observed[sector_name][skill_id] += 1

        return self.sector_skill_observed

    def get_observed_skills_for_occupation(self, occ_id: str, resolve_labels: bool = False, top_k: int = 20):
        """
                Return the observed skills for a single occupation, sorted by count.

                Parameters
                ----------
                occ_id : str
                    Occupation id
                resolve_labels : bool
                    If True, tries to resolve skill labels using self.skill_map
                top_k : int
                    Maximum number of returned skills

                Returns
                -------
                List[dict]
                """
        occ_id = str(occ_id).strip()
        counter = self.occ_skill_observed.get(occ_id, Counter())
        total_skill_mentions = sum(counter.values())

        results = []
        for skill_id, count in counter.most_common(top_k):
            entry = {
                "skill_id": skill_id,
                "count": count,
                "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0,
            }
            if resolve_labels:
                meta = self.skill_map.get(skill_id, {})
                entry["label"] = meta.get("label", skill_id)
                entry["is_green"] = meta.get("is_green", False)
                entry["is_digital"] = meta.get("is_digital", False)
            results.append(entry)

        return results


    def summarize_observed_sector_skills(
        self,
        sector_level: str = "isco_group",
        resolve_labels: bool = False,
        top_k: int = 20
    ):
        """
        Build a readable summary of observed skills per sector, starting from
        self.sector_skill_observed.

        Parameters
        ----------
        sector_level : str
            Informational only for now; kept for consistency with previous steps
        resolve_labels : bool
            If True, resolves skill labels through self.skill_map
        top_k : int
            Max number of top skills to return per sector

        Returns
        -------
        List[dict]
            One entry per sector
        """
        results = []

        for sector_name, counter in self.sector_skill_observed.items():
            total_skill_mentions = sum(counter.values())

            top_skills = []
            for skill_id, count in counter.most_common(top_k):
                entry = {
                    "skill_id": skill_id,
                    "count": count,
                    "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0
                }

                if resolve_labels:
                    meta = self.skill_map.get(skill_id, {})
                    entry["label"] = meta.get("label", skill_id)
                    entry["is_green"] = meta.get("is_green", False)
                    entry["is_digital"] = meta.get("is_digital", False)

                top_skills.append(entry)

            results.append({
                "sector": sector_name,
                "total_skill_mentions": total_skill_mentions,
                "unique_skills": len(counter),
                "top_skills": top_skills
            })

        return sorted(results, key=lambda x: x["total_skill_mentions"], reverse=True)

    def build_and_summarize_observed_sector_skills(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        resolve_labels: bool = False,
        top_k: int = 20,
        reset: bool = True
    ):
        """
        Convenience method:
        1. builds sector -> skill observed matrix
        2. returns a readable summary
        """
        self.build_observed_sector_skill_matrix(
            jobs=jobs,
            sector_level=sector_level,
            reset=reset
        )

        return self.summarize_observed_sector_skills(
            sector_level=sector_level,
            resolve_labels=resolve_labels,
            top_k=top_k
        )

    def build_and_summarize_observed_sector_skills(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        resolve_labels: bool = False,
        top_k: int = 20,
        reset: bool = True
    ):
        """
        Convenience method:
        1. builds sector -> skill observed matrix
        2. returns a readable summary
        """
        self.build_observed_sector_skill_matrix(
            jobs=jobs,
            sector_level=sector_level,
            reset=reset
        )

        return self.summarize_observed_sector_skills(
            sector_level=sector_level,
            resolve_labels=resolve_labels,
            top_k=top_k
        )
    def build_and_summarize_canonical_sector_skills(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        resolve_labels: bool = False,
        top_k: int = 20,
        reset: bool = True
    ):
        """
        Convenience method:
        1. builds sector -> canonical skill matrix
        2. returns a readable summary
        """
        self.build_canonical_sector_skill_matrix(
            jobs=jobs,
            sector_level=sector_level,
            reset=reset
        )

        return self.summarize_canonical_sector_skills(
            resolve_labels=resolve_labels,
            top_k=top_k
        )
    def summarize_single_sector(
        self,
        sector_name: str,
        resolve_labels: bool = False,
        top_k: int = 20
    ):
        """
        Return a readable summary for one specific sector only.
        """
        from collections import Counter  # sicurezza se non già importato

        sector_name = str(sector_name).strip()
        counter = self.sector_skill_observed.get(sector_name, Counter())
        total_skill_mentions = sum(counter.values())

        top_skills = []
        for skill_id, count in counter.most_common(top_k):
            entry = {
                "skill_id": skill_id,
                "count": count,
                "frequency": round(count / total_skill_mentions, 6) if total_skill_mentions > 0 else 0.0
            }

            if resolve_labels:
                meta = self.skill_map.get(skill_id, {})
                entry["label"] = meta.get("label", skill_id)
                entry["is_green"] = meta.get("is_green", False)
                entry["is_digital"] = meta.get("is_digital", False)

            top_skills.append(entry)

        return {
            "sector": sector_name,
            "total_skill_mentions": total_skill_mentions,
            "unique_skills": len(counter),
            "top_skills": top_skills
        }
    def load_local_esco_support(self):
        """
        Load local CSV support files for:
        - occupation metadata
        - skill hierarchy
        - occupation-skill canonical relations

        This step is safe and incremental: if a file is missing, it is skipped.
        """
        base_dir = os.getcwd()

        occupations_file = os.path.join(base_dir, "complementary_data", "occupations_en.csv")
        skills_hierarchy_file = os.path.join(base_dir, "complementary_data", "skillsHierarchy_en.csv")
        occ_skill_rel_file = os.path.join(base_dir, "complementary_data", "occupationSkillRelations_en.csv")
        isco_groups_file = os.path.join(base_dir, "complementary_data", "ISCOGroups_en.csv")
        skill_groups_file = os.path.join(base_dir, "complementary_data", "skillGroups_en.csv")

        # 1) Occupation metadata
        if os.path.exists(occupations_file):
            try:
                with open(occupations_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        occ_id = (
                            row.get("conceptUri")
                            or row.get("id")
                            or row.get("occupationUri")
                            or ""
                        ).strip()
                        if not occ_id:
                            continue

                        self.occupation_meta[occ_id] = {
                            "label": (row.get("preferredLabel") or row.get("label") or "").strip(),
                            "isco_group": (row.get("iscoGroup") or "").strip(),
                            "nace_code": (row.get("naceCode") or "").strip(),
                            "raw": row,
                        }
                logger.info(f"Loaded occupation metadata: {len(self.occupation_meta)}")
            except Exception as e:
                logger.warning(f"Could not load occupations_lt.csv: {e}")

        # 2) Skill hierarchy
        if os.path.exists(skills_hierarchy_file):
            try:
                with open(skills_hierarchy_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        skill_id = (
                                row.get("conceptUri")
                                or row.get("id")
                                or row.get("skillUri")
                                or ""
                        ).strip()
                        if not skill_id:
                            continue

                        level_1 = (row.get("level1") or row.get("ESCO Level 1") or row.get("skillLevel1") or "").strip()
                        level_2 = (row.get("level2") or row.get("ESCO Level 2") or row.get("skillLevel2") or "").strip()
                        level_3 = (row.get("level3") or row.get("ESCO Level 3") or row.get("skillLevel3") or "").strip()

                        self.skill_hierarchy[skill_id] = {
                            "level_1": level_1,
                            "level_2": level_2,
                            "level_3": level_3,
                            "raw": row,
                        }

                logger.info(f"Loaded skill hierarchy: {len(self.skill_hierarchy)}")
            except Exception as e:
                logger.warning(f"Could not load skillsHierarchy_en.csv: {e}")
        # 3) Skill group labels
        if os.path.exists(skill_groups_file):
            try:
                with open(skill_groups_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        gid = (
                            row.get("conceptUri")
                            or row.get("id")
                            or row.get("skillGroup")
                            or ""
                        ).strip()

                        label = (
                            row.get("preferredLabel")
                            or row.get("label")
                            or row.get("title")
                            or ""
                        ).strip()

                        if gid and label:
                            self.skill_group_labels[gid] = label

                            # also store short-code fallback, e.g. .../skill/S4.8 -> S4.8
                            short_gid = gid.rstrip("/").split("/")[-1]
                            if short_gid:
                                self.skill_group_labels[short_gid] = label

                logger.info(f"Loaded skill group labels: {len(self.skill_group_labels)}")
            except Exception as e:
                logger.warning(f"Could not load skillGroups_en.csv: {e}")

        # 4) Occupation-skill canonical relations
        if os.path.exists(occ_skill_rel_file):
            try:
                with open(occ_skill_rel_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        occ_id = (
                            row.get("occupationUri")
                            or row.get("occupation")
                            or row.get("occupation_id")
                            or row.get("conceptUriOccupation")
                            or ""
                        ).strip()

                        skill_id = (
                            row.get("skillUri")
                            or row.get("skill")
                            or row.get("skill_id")
                            or row.get("conceptUriSkill")
                            or ""
                        ).strip()

                        if occ_id and skill_id:
                            # save for each occupations the skill requested
                            self.occ_skill_relations[occ_id].add(skill_id)
                logger.info(f"Loaded canonical occupation-skill relations: {len(self.occ_skill_relations)} occupations")
            except Exception as e:
                logger.warning(f"Could not load occupationSkillRelations_lt.csv: {e}")

        # 4) Optional ISCO group labels
        if os.path.exists(isco_groups_file):
            try:
                with open(isco_groups_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        raw_gid = (
                                row.get("conceptUri")
                                or row.get("id")
                                or row.get("iscoGroup")
                                or ""
                        ).strip()

                        gid = raw_gid.split("/")[-1].replace("C", "")

                        label = (row.get("preferredLabel") or row.get("label") or "").strip()
                        if gid:
                            self.occupation_group_labels[gid] = label
                logger.info(f"Loaded ISCO group labels: {len(self.occupation_group_labels)}")
            except Exception as e:
                logger.warning(f"Could not load ISCOGroups_lt.csv: {e}")

    def get_skill_group(self, skill_id: str, level: int = 2) -> str:
        """
        Resolve the ESCO skill group for a skill using the local skill hierarchy.
        Returns a normalized short group id when possible (e.g. S4.8 instead of full URI).
        """
        skill_id = str(skill_id).strip()
        if not skill_id:
            return "Skill group not specified"

        meta = self.skill_hierarchy.get(skill_id)
        if meta:
            if level == 1:
                group_id = (meta.get("level_1") or "").strip()
            elif level == 3:
                group_id = (meta.get("level_3") or "").strip()
            else:
                group_id = (meta.get("level_2") or "").strip()

            if group_id:
                # normalize URI -> short code
                if group_id.startswith("http"):
                    return group_id.rstrip("/").split("/")[-1]
                return group_id

        skill_meta = self.skill_map.get(skill_id, {})
        label = skill_meta.get("label", "").strip()
        if label:
            return label

        return "Skill group not specified"


    def build_sectoral_intelligence(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 1,
        occupation_level: int = 1,
        resolve_labels: bool = False,
        top_k_skills: int = 10,
        top_k_groups: int = 10,
        reset: bool = True
    ):
        """
        Build a unified sectoral intelligence payload combining:
        - observed skills
        - canonical skills
        - observed skill groups
        - canonical skill groups
        - official matrix skill groups
        """
        # 1. Build all required layers
        self.build_observed_sector_skill_matrix(
            jobs=jobs,
            sector_level=sector_level,
            reset=reset
        )

        self.build_canonical_sector_skill_matrix(
            jobs=jobs,
            sector_level=sector_level,
            reset=reset
        )

        self.build_observed_sector_skillgroup_matrix(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            reset=reset
        )

        self.build_canonical_sector_skillgroup_matrix(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            reset=reset
        )

        self.build_official_matrix_sector_skillgroup_profile(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            occupation_level=occupation_level,
            reset=reset
        )

        # 2. Collect all sector names seen in any layer
        sectors = set()
        sectors.update(self.sector_skill_observed.keys())
        sectors.update(self.sector_skill_canonical.keys())
        sectors.update(self.sector_skillgroup_observed.keys())
        sectors.update(self.sector_skillgroup_canonical.keys())
        sectors.update(self.matrix_profiles.keys())

        results = []

        for sector_name in sorted(sectors):
            observed_skills = self.summarize_single_sector(
                sector_name=sector_name,
                resolve_labels=resolve_labels,
                top_k=top_k_skills
            )

            canonical_skills = {
                "sector": sector_name,
                "top_skills": self.get_canonical_skills_for_sector(
                    sector_name=sector_name,
                    resolve_labels=resolve_labels,
                    top_k=top_k_skills
                ),
                "total_skill_mentions": sum(self.sector_skill_canonical.get(sector_name, Counter()).values()),
                "unique_skills": len(self.sector_skill_canonical.get(sector_name, Counter()))
            }

            group_profiles = self.compare_all_group_profiles_for_sector(
                sector_name=sector_name,
                top_k=top_k_groups
            )

            results.append({
                "sector": sector_name,
                "sector_label": self.get_sector_label(sector_name),

                "observed_skills": observed_skills,
                "canonical_skills": canonical_skills,

                "observed_groups": group_profiles["observed_groups"],
                "canonical_groups": group_profiles["canonical_groups"],

                "matrix_groups": group_profiles["official_matrix_groups"]
            })

        return results
    def build_single_sector_intelligence(
        self,
        sector_name: str,
        resolve_labels: bool = False,
        top_k_skills: int = 10,
        top_k_groups: int = 10
    ):
        """
        Return the unified sectoral intelligence block for a single already-built sector.
        Assumes the matrices have already been populated.
        """
        sector_name = str(sector_name).strip()

        observed_skills = self.summarize_single_sector(
            sector_name=sector_name,
            resolve_labels=resolve_labels,
            top_k=top_k_skills
        )

        canonical_skills = {
            "sector": sector_name,
            "top_skills": self.get_canonical_skills_for_sector(
                sector_name=sector_name,
                resolve_labels=resolve_labels,
                top_k=top_k_skills
            ),
            "total_skill_mentions": sum(self.sector_skill_canonical.get(sector_name, Counter()).values()),
            "unique_skills": len(self.sector_skill_canonical.get(sector_name, Counter()))
        }

        group_profiles = self.compare_all_group_profiles_for_sector(
            sector_name=sector_name,
            top_k=top_k_groups
        )

        return {
            "sector": sector_name,
            "sector_label": self.get_sector_label(sector_name),

            "observed_skills": observed_skills,
            "canonical_skills": canonical_skills,

            "observed_groups": group_profiles["observed_groups"],
            "canonical_groups": group_profiles["canonical_groups"],

            "matrix_groups": group_profiles["official_matrix_groups"]
        }
    def get_sector_from_occupation(self, occ_id: str, level: str = "isco_group") -> str:
        """
        Resolve a sector/group label from an occupation id.

        Resolution order:
        1. Local occupation metadata loaded from occupations_lt.csv
        2. Tracker-derived sector_map fallback
        3. Generic fallback

        Parameters
        ----------
        occ_id : str
            ESCO occupation URI or id
        level : str
            Currently supported:
            - "isco_group": uses local isco_group if available
            - "label": returns the occupation label
            - "nace_code": returns the NACE code if available

        Returns
        -------
        str
            A readable sector/group string


        1. CSV locale (migliore qualità)
            ├─ NACE (se richiesto)
            ├─ Label (se richiesto)
            └─ ISCO group (default)

        2. Tracker API fallback
           └─ sector_map

        3. Default
           └─ "Sector not specified"
        """
        occ_id = str(occ_id).strip()
        if not occ_id:
            return "Sector not specified"

        # 1) Local metadata preferred
        meta = self.occupation_meta.get(occ_id)
        if meta:
            if level == "nace_code":
                nace = meta.get("nace_code", "").strip()
                if nace:
                    return nace

            if level == "label":
                label = meta.get("label", "").strip()
                if label:
                    return label

            # default: ISCO group if available
            isco_group = meta.get("isco_group", "").strip()
            if isco_group:
                # if we have a readable label for the ISCO group, use it
                return self.occupation_group_labels.get(isco_group, isco_group)

            # fallback to occupation label
            label = meta.get("label", "").strip()
            if label:
                return label

        # 2) Tracker-derived fallback
        tracker_label = self.sector_map.get(occ_id, "").strip()
        if tracker_label:
            return tracker_label

        # 3) Last-resort fallback
        return "Sector not specified"

    def get_primary_occupation_id(self, job: dict) -> str:
        """
        Extract the primary occupation id from a job in a backward-compatible way.
        """
        job_occs = job.get("occupations", [])
        legacy_occ = job.get("occupation_id")

        if job_occs and len(job_occs) > 0:
            return str(job_occs[0]).strip()

        if legacy_occ:
            return str(legacy_occ).strip()

        return ""

    def get_occupation_ids(self, job: dict) -> List[str]:
        """
        Extract all occupation ids from a job in a backward-compatible way.
        """
        occ_ids = []

        job_occs = job.get("occupations", [])
        legacy_occ = job.get("occupation_id")

        for occ in job_occs:
            occ = str(occ).strip()
            if occ:
                occ_ids.append(occ)

        if legacy_occ:
            legacy_occ = str(legacy_occ).strip()
            if legacy_occ and legacy_occ not in occ_ids:
                occ_ids.append(legacy_occ)

        return occ_ids



    def _load_skill_uris_from_csv(self, path: str) -> set[str]:
        uris = set()
        if not os.path.exists(path):
            logger.warning(f"CSV non trovato: {path}")
            return uris

        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                uri = (row.get("conceptUri") or "").strip()
                if uri:
                    uris.add(uri)

        logger.info(f"Caricati {len(uris)} URI da {path}")
        return uris


    def request_stop(self):
        """
           Requests a cooperative stop of all ongoing operations.

           This method sets an internal flag (`stop_requested`) that is periodically
           checked by long-running asynchronous processes (e.g., data fetching,
           skill translation, analysis loops). When the flag is detected, the engine
           gracefully interrupts execution at the next safe checkpoint.

           This avoids abrupt termination and ensures partial results can still be returned.

           Side Effects:
               - Sets `self.stop_requested = True`
           """
        self.stop_requested = True
        logger.warning("!!! STOP RILEVATO !!! Interruzione programmata dei processi.")

    async def _get_token(self):
        """
            Authenticates with the external Tracker API and retrieves an access token.

            The token is required for all subsequent API calls (jobs, skills, occupations).
            It is stored internally and reused until expiration.

            Returns:
                Optional[str]: Bearer token string if authentication succeeds,
                               None if authentication fails.

            External Dependencies:
                - POST {TRACKER_API}/login

            Failure Handling:
                - Logs error and returns None without raising exception.
            """
        try:
            resp = await self.client.post(
                f"{self.api_url}/login",
                json={"username": self.username, "password": self.password}
            )
            self.token = resp.text.replace('"', '')
            return self.token
        except Exception as e:
            logger.error(f"Errore Login: {e}")
            return None

    async def fetch_occupation_labels(self, occ_uris: List[str], page_size: int = 500):
        """
           Resolves occupation URIs into human-readable sector labels.

           This method enriches job data by mapping ESCO occupation identifiers
           to their corresponding sector names. The results are cached in `self.sector_map`
           to avoid redundant API calls.

           Args:
               occ_uris (List[str]): List of occupation identifiers (ESCO URIs).
               page_size (int): Pagination size for API requests.

           Behavior:
               - Filters out already known URIs using internal cache.
               - Fetches data in batches (size=40).
               - Updates `self.sector_map` with {occupation_id: label}.

           External Dependencies:
               - POST {TRACKER_API}/occupations

           Side Effects:
               - Modifies `self.sector_map`

           Early Exit:
               - Returns immediately if `stop_requested` is True.
           """

        uris = [str(u).strip() for u in occ_uris if u and str(u).strip() not in self.sector_map]

        if not uris or self.stop_requested: return

        if not self.token: await self._get_token()

        batch_size = 40
        for i in range(0, len(uris), batch_size):
            if self.stop_requested: break
            batch = uris[i:i + batch_size]
            try:
                res = await self.client.post(
                    f"{self.api_url}/occupations",
                    headers={"Authorization": f"Bearer {self.token}"},
                    data={"ids": batch}
                )
                if res.status_code == 200:
                    for o in res.json().get("items", []):
                        # Salviamo l'ID e la label (Preferred Label)
                        self.sector_map[str(o.get("id")).strip()] = str(o.get("label", ""))
            except:
                continue

    def get_regional_projections(self, jobs: List[dict], demo: bool = False):
        """
           Computes geographical projections of job data across NUTS hierarchy levels.

           This method aggregates job postings by location and derives regional
           intelligence using both raw location codes and hierarchical NUTS levels
           (NUTS1, NUTS2, NUTS3).

           It also calculates the Location Quotient (LQ) to measure skill specialization.

           Args:
               jobs (List[dict]): List of job postings.
               demo (bool): If True, simulates NUTS-level granularity when only
                            national codes are available.

           Returns:
               dict:
                   {
                       "raw": [...],
                       "nuts1": [...],
                       "nuts2": [...],
                       "nuts3": [...]
                   }

           Key Concepts:
               - Market Share: % of total jobs in a region
               - Specialization (LQ): relative concentration of a skill in a region

           Behavior:
               - Aggregates job counts per region
               - Aggregates skill counts per region
               - Computes LQ per skill
               - Returns top skills per region

           Note:
               In demo mode, NUTS codes are synthetically generated using index-based distribution.
           """
        raw_map = {}
        nuts_map = {"NUTS1": {}, "NUTS2": {}, "NUTS3": {}}
        global_counts = {}
        total_jobs = len(jobs) if jobs else 1
        for idx, job in enumerate(jobs):
            # Prendiamo il location_code originale (es. "IT", "SE", "FR")
            loc_original = str(job.get("location_code", "EU")).strip()

            # Estraiamo il prefisso nazione (primi 2 caratteri)

            # Se il codice è solo nazionale (lungo 2), generiamo un NUTS3 dinamico
            if demo and len(loc_original) <= 2:
                country_prefix = loc_original[:2].upper() if len(loc_original) >= 2 else "EU"

                # Creiamo una scomposizione "pseudo-reale" usando l'indice
                # NUTS structure: [CC][Level1][Level2][Level3] -> ES: IT 1 2 1
                l1 = (idx % 3) + 1  # Varia tra 1 e 3
                l2 = (idx % 4) + 1  # Varia tra 1 e 4
                l3 = (idx % 5)  # Varia tra 0 e 4
                loc_projected = f"{country_prefix}{l1}{l2}{l3}"
            else:
                loc_projected = loc_original

            # -----------------------

            nuts_levels = {
                "NUTS1": loc_projected[:3],  # Es: ITC
                "NUTS2": loc_projected[:4] if len(loc_projected) >= 4 else None,  # Es: ITC4
                "NUTS3": loc_projected if len(loc_projected) >= 5 else None  # Es: ITC4C

            }

            # 2. INCREMENTO JOB COUNT (Una sola volta per job!)
            # Strategia RAW
            if loc_original not in raw_map: raw_map[loc_original] = {"count": 0, "skills": {}}
            raw_map[loc_original]["count"] += 1

            # Strategia NUTS
            for level, code in nuts_levels.items():
                if code:
                    if code not in nuts_map[level]: nuts_map[level][code] = {"count": 0, "skills": {}}
                    nuts_map[level][code]["count"] += 1

            # 3. AGGREGAZIONE SKILLS
            for s_uri in job.get("skills", []):
                label = self.skill_map.get(s_uri, {}).get("label", s_uri)

                # Update globale per LQ
                global_counts[label] = global_counts.get(label, 0) + 1

                # Update RAW
                raw_map[loc_original]["skills"][label] = raw_map[loc_original]["skills"].get(label, 0) + 1

                # Update NUTS
                for level, code in nuts_levels.items():
                    if code:
                        node_skills = nuts_map[level][code]["skills"]
                        node_skills[label] = node_skills.get(label, 0) + 1

        # --- FORMATTAZIONE FINALE (Invariata) ---
        def format_output(source_map):
            formatted = []
            for code, data in source_map.items():
                skills_list = []
                for s_name, count in data["skills"].items():
                    # Calcolo Location Quotient (Specializzazione)
                    lq = (count / data["count"]) / (global_counts[s_name] / total_jobs)
                    skills_list.append({
                        "skill": s_name,
                        "count": count,
                        "specialization": round(lq, 2)
                    })
                formatted.append({
                    "code": code,
                    "total_jobs": data["count"],
                    "market_share": round((data["count"] / total_jobs) * 100, 2),
                    "top_skills": sorted(skills_list, key=lambda x: x["count"], reverse=True)[:10]
                })
            return sorted(formatted, key=lambda x: x["total_jobs"], reverse=True)

        return {
            "raw": format_output(raw_map),
            "nuts1": format_output(nuts_map["NUTS1"]),
            "nuts2": format_output(nuts_map["NUTS2"]),
            "nuts3": format_output(nuts_map["NUTS3"])
        }

    async def fetch_all_jobs(self, filters: dict, page_size: int = 500):
        """
           Fetches all job postings from the Tracker API using pagination and caching.

           This method orchestrates the retrieval of job data based on user-defined filters.
           It supports persistent caching to avoid redundant API calls for identical queries.

           Args:
               filters (dict): Query parameters (keywords, date range, etc.).
               page_size (int): Number of records per API request.

           Returns:
               List[dict]: List of job postings retrieved from the API or cache.

           Behavior:
               - Generates a hash signature for the query
               - Checks disk cache (`cache_data/`)
               - If cache miss:
                   - Fetches paginated results from API
                   - Stores results in cache
               - Supports cooperative stop via `stop_requested`

           External Dependencies:
               - POST {TRACKER_API}/jobs

           Failure Handling:
               - Handles ReadTimeout gracefully
               - Logs errors and returns partial results if interrupted

           Side Effects:
               - Writes cache files to disk
           """
        # Non resettiamo stop_requested qui, lo facciamo negli endpoint all'inizio
        query_sig = hashlib.md5(json.dumps(filters, sort_keys=True).encode()).hexdigest()
        cache_dir, cache_file = "cache_data", f"cache_data/search_{query_sig}.json"

        if os.path.exists(cache_file):
            logger.info(f"Cache Hit: {query_sig}")
            with open(cache_file, 'r') as f: return json.load(f)

        if not self.token: await self._get_token()

        all_jobs, page = [], 1
        headers = {"Authorization": f"Bearer {self.token}"}

        while True:
            if self.stop_requested:
                logger.warning("Fetch fermato per stop richiesto.")
                break

            try:
                res = await self.client.post(
                    f"{self.api_url}/jobs",
                    headers=headers,
                    data=filters,
                    params={"page": page, "page_size": page_size}
                )

                if res.status_code != 200: break
                data = res.json()
                items = data.get("items", [])
                all_jobs.extend(items)

                total = data.get("count", 0)
                logger.info(f"Fetching: {len(all_jobs)}/{total} (Pagina {page})")

                if len(all_jobs) >= total or not items: break
                page += 1
                await asyncio.sleep(0.01)  # Checkpoint per event loop

            except httpx.ReadTimeout:
                logger.error("Timeout durante il fetch. L'API Tracker è lenta.")
                break
            except Exception as e:
                logger.error(f"Errore fetch: {e}")
                break

        if not self.stop_requested and all_jobs:
            if not os.path.exists(cache_dir): os.makedirs(cache_dir)
            with open(cache_file, 'w') as f:
                json.dump(all_jobs, f)

        return all_jobs

    async def fetch_skill_names(self, skill_uris: List[str], page_size: int = 500):
        """
            Resolves skill URIs into enriched skill metadata (label + Twin Transition tags).

            This method translates ESCO skill identifiers into human-readable labels
            and assigns semantic tags for:
                - Digital skills
                - Green skills

            Args:
                skill_uris (List[str]): List of skill identifiers.
                page_size (int): Pagination size for API requests.

            Behavior:
                - Filters already known skills using cache
                - Fetches data in batches
                - Applies heuristic keyword matching to classify:
                    - is_digital
                    - is_green
                - Stores results in `self.skill_map`

            External Dependencies:
                - POST {TRACKER_API}/skills

            Side Effects:
                - Modifies `self.skill_map`

            Early Exit:
                - Returns immediately if `stop_requested` is True.
            """
        uris = [u for u in skill_uris if u not in self.skill_map]
        if not uris or self.stop_requested: return

        if not self.token:
            await self._get_token()
        # TODO: Placeholder for Twin Transition tagging (Task 3.5 requirement)
        GREEN_KEYWORDS = {
            "sustainable", "sustainable", "ecology", "circular", "carbon", "renewable",
            "energy", "photovoltaic", "recycling", "environmental", "climate", "efficiency"
        }
        DIGITAL_KEYWORDS = {
            "software", "digital", "ai", "artificial intelligence", "coding", "cloud",
            "data", "computing", "cybersecurity", "web", "automation", "programming"
        }
        batch_size = 40
        for i in range(0, len(uris), batch_size):
            if self.stop_requested: break
            batch = uris[i:i + batch_size]
            try:
                res = await self.client.post(
                    f"{self.api_url}/skills",
                    headers={"Authorization": f"Bearer {self.token}"},
                    data={"ids": batch, "keywords_logic": "or"},
                    params={"page": 1, "page_size": page_size}
                )
                if res.status_code == 200:
                    for s in res.json().get("items", []):
                        s_id = s.get("id")
                        label = s.get("label")

                        # Intelligence: Tagging Twin Transition (Logica euristica basata su metadati o URI)
                        # Nota: In produzione qui interrogheremmo i metadati ESCO
                        # TODO: eliminare is_green, is_digital
                        is_green = False
                        is_digital = False

                        self.skill_map[s_id] = {
                            "label": label,
                            "is_green": is_green,
                            "is_digital": is_digital
                        }
            except:
                continue

    async def analyze_market_data(self, raw_jobs: List[dict]):
        """
           Performs multi-dimensional aggregation of job market data.

           This method is the core analytical engine of the Projector. It processes
           raw job postings and extracts structured intelligence across multiple dimensions:
               - Skills demand
               - Employers
               - Job titles
               - Geographic distribution
               - Sector distribution

           Args:
               raw_jobs (List[dict]): Raw job postings retrieved from the Tracker API.

           Returns:
               dict:
                   {
                       "total_jobs": int,
                       "rankings": {
                           "skills": [...],
                           "employers": [...],
                           "job_titles": [...],
                           "locations": [...],
                           "sectors": [...]
                       }
                   }

           Behavior:
               - Uses Counter-based aggregation for performance
               - Builds skill-to-sector relationships
               - Enriches skills using `fetch_skill_names`
               - Computes frequency and sector spread

           Performance Notes:
               - Includes async checkpoints to avoid blocking event loop
               - Handles large datasets (40k+ jobs)

           Early Exit:
               - Stops processing if `stop_requested` is True
           """
        if not raw_jobs: return self._empty_res()

        s_cnt, e_cnt, t_cnt, l_cnt, sec_cnt = Counter(), Counter(), Counter(), Counter(), Counter()
        skill_sector_map = {}  # La matrice Skill -> [Settori]

        for i, job in enumerate(raw_jobs):
            if self.stop_requested: break
            if i > 0 and i % 2000 == 0: await asyncio.sleep(0)

            # 1. Occupation -> sector
            occ_ids = self.get_occupation_ids(job)

            for occ_id in occ_ids:
                sector_name = self.get_sector_from_occupation(occ_id, level="isco_group")
                sec_cnt[sector_name] += 1

            for s_uri in job.get("skills", []):
                s_uri = str(s_uri).strip()
                s_cnt[s_uri] += 1
                if s_uri not in skill_sector_map:
                    skill_sector_map[s_uri] = Counter()

                for occ_id in occ_ids:
                    sector_name = self.get_sector_from_occupation(occ_id, level="isco_group")
                    skill_sector_map[s_uri][sector_name] += 1

            e_cnt[job.get("organization_name") or "N/D"] += 1
            t_cnt[job.get("title") or "N/D"] += 1
            l_cnt[job.get("location_code") or "N/D"] += 1

        # 4. Arricchimento nomi (Skill + Settori)
        await self.fetch_skill_names(list(s_cnt.keys()))

        # 5. Output con Intelligence
        enriched_ranking = []
        for k, v in s_cnt.most_common():
            skill_info = self.skill_map.get(k, {"label": k.split('/')[-1], "is_green": False, "is_digital": False})

            # Intelligence: Calcolo trasversalità (Phase 1 core)
            sectors_involved = skill_sector_map[k]

            enriched_ranking.append({
                "name": skill_info["label"],
                "frequency": v,
                "skill_id": k,
                "is_green": skill_info["is_green"],
                "is_digital": skill_info["is_digital"],
                "sector_spread": len(sectors_involved),  # Quanti settori la cercano?
                "primary_sector": sectors_involved.most_common(1)[0][0] if sectors_involved else "N/D"
            })

        return {
            "total_jobs": len(raw_jobs),
            "rankings": {
                "skills": enriched_ranking,
                "employers": [{"name": k, "count": v} for k, v in e_cnt.most_common(10)],
                "job_titles": [{"name": k, "count": v} for k, v in t_cnt.most_common(10)],
                "sectors": [{"name": k, "count": v} for k, v in sec_cnt.most_common(10)]  # NUOVO
            },
            "geo": [{"location": k, "job_count": v} for k, v in l_cnt.most_common()]
        }

    def _empty_res(self):
        return {
            "total_jobs": 0,
            "rankings": {
                "skills": [],
                "employers": [],
                "job_titles": [],
                "sectors": []  # Consistenza Phase 1
            },
            "geo": []
        }

    def _empty_insights_p1(self):
        return {
            "ranking": [],
            "sectors": [],
            "job_titles": [],
            "employers": [],
            "trends": {
                "market_health": {
                    "status": "stable",
                    "volume_growth_percentage": 0.0
                },
                "trends": []
            },
            "regional": {
                "raw": [],
                "nuts1": [],
                "nuts2": [],
                "nuts3": []
            },
            "sectoral": None
        }

    # --- METODO PRIVATO CONDIVISO (IL CERVELLO) ---
    def _compare_periods(self, res_a, res_b):
        """Calcola i delta e arricchisce con Intelligence (Phase 1)."""
        dict_a = {s["skill_id"]: s for s in res_a["rankings"]["skills"]}
        dict_b = {s["skill_id"]: s for s in res_b["rankings"]["skills"]}
        trends = []

        for s_id in set(list(dict_a.keys()) + list(dict_b.keys())):
            v_a = dict_a.get(s_id, {}).get("frequency", 0)
            info_b = dict_b.get(s_id, {})
            v_b = info_b.get("frequency", 0)

            name = info_b.get("name") or dict_a.get(s_id, {}).get("name")
            primary_sector = info_b.get("primary_sector", "N/D")

            if v_a == 0:
                growth, t_type = "new_entry", "emerging"
            elif v_b == 0:
                growth, t_type = -100.0, "declining"
            else:
                growth = round(((v_b - v_a) / v_a) * 100, 2)
                t_type = "emerging" if growth > 0 else "declining" if growth < 0 else "stable"

            trends.append({
                "name": name,
                "growth": growth,
                "trend_type": t_type,
                "primary_sector": primary_sector,
                "is_green": info_b.get("is_green", False),
                "is_digital": info_b.get("is_digital", False)
            })

        trends.sort(key=lambda x: float('inf') if x["growth"] == "new_entry" else x["growth"], reverse=True)
        vol_growth = round(((res_b["total_jobs"] - res_a["total_jobs"]) / res_a["total_jobs"] * 100), 2) if res_a[
                                                                                                                "total_jobs"] > 0 else 0

        return {
            "market_health": {
                "status": "expanding" if vol_growth > 0 else "shrinking",
                "volume_growth_percentage": vol_growth
            },
            "trends": trends
        }

    # --- METODO 1: OTTIMIZZATO (IN-MEMORY) ---
    async def calculate_trends_from_data(self, all_jobs: List[dict], min_date: str, max_date: str):
        mid = self._get_midpoint(min_date, max_date)
        jobs_a = [j for j in all_jobs if j.get("upload_date", "") <= mid]
        jobs_b = [j for j in all_jobs if j.get("upload_date", "") > mid]

        res_a = await self.analyze_market_data(jobs_a)
        res_b = await self.analyze_market_data(jobs_b)
        return self._compare_periods(res_a, res_b)

    # --- METODO 2: STANDALONE (CON FETCH) ---
    async def calculate_smart_trends(self, base_filters: dict, min_date: str, max_date: str):
        mid = self._get_midpoint(min_date, max_date)
        f_a = {**base_filters, "min_upload_date": min_date, "max_upload_date": mid}
        f_b = {**base_filters, "min_upload_date": mid, "max_upload_date": max_date}  # Semplificato per brevità

        res_a = await self.analyze_market_data(await self.fetch_all_jobs(f_a))
        if self.stop_requested: return self._stop_trend_res()

        res_b = await self.analyze_market_data(await self.fetch_all_jobs(f_b))
        return self._compare_periods(res_a, res_b)

    def _get_midpoint(self, d1, d2):
        dt1, dt2 = datetime.strptime(d1, "%Y-%m-%d"), datetime.strptime(d2, "%Y-%m-%d")
        return (dt1 + timedelta(days=(dt2 - dt1).days // 2)).strftime("%Y-%m-%d")
    def get_esco_matrix_sheet_name(self, skill_group_level: int = 1, occupation_level: int = 1) -> str:
        """
        Resolve the official ESCO matrix sheet name.

        Example:
        skill_group_level=1, occupation_level=1 -> Matrix 1.1
        """
        return f"Matrix {skill_group_level}.{occupation_level}"
    def get_occupation_group_id_for_matrix(self, occ_id: str, occupation_level: int = 1) -> str:
        """
        Resolve the occupation group id used by the official ESCO matrix.

        Current incremental version:
        - read local occupation_meta[occ_id]["isco_group"]
        - try to reduce it to the requested ISCO level if it looks like a C-code

        Examples:
        - occupation_level=1 -> C2
        - occupation_level=2 -> C25
        - occupation_level=3 -> C251
        - occupation_level=4 -> C2512
        """
        occ_id = str(occ_id).strip()
        if not occ_id:
            return ""

        meta = self.occupation_meta.get(occ_id, {})
        group_id = str(meta.get("isco_group", "")).strip()

        if not group_id:
            return ""

        # if already in ESCO/ISCO URI form like http://data.europa.eu/esco/isco/C2512
        if group_id.startswith("http"):
            code = group_id.rstrip("/").split("/")[-1]
        else:
            code = group_id

        code = code.strip()
        if code and not code.startswith("C") and code[0].isdigit():
            code = f"C{code}"
        # normalize to requested level if possible
        if code.startswith("C"):
            if occupation_level == 1:
                return f"C{code[1:2]}" if len(code) >= 2 else code
            elif occupation_level == 2:
                return f"C{code[1:3]}" if len(code) >= 3 else code
            elif occupation_level == 3:
                return f"C{code[1:4]}" if len(code) >= 4 else code
            elif occupation_level == 4:
                return f"C{code[1:5]}" if len(code) >= 5 else code

        return code
    def get_official_esco_profile_for_occupation(
        self,
        occ_id: str,
        skill_group_level: int = 1,
        occupation_level: int = 1
    ):
        """
        Return the official ESCO matrix profile for a given occupation, by:
        1. resolving the occupation group id
        2. selecting the proper matrix sheet
        3. returning the stored profile
        """
        group_code = self.get_occupation_group_id_for_matrix(
            occ_id=occ_id,
            occupation_level=occupation_level
        )
        if not group_code:
            return None

        sheet_name = self.get_esco_matrix_sheet_name(
            skill_group_level=skill_group_level,
            occupation_level=occupation_level
        )

        uri_key = f"http://data.europa.eu/esco/isco/{group_code}"

        profile = self.esco_matrix_profiles.get((sheet_name, uri_key))
        if profile:
            return {
                "sheet_name": sheet_name,
                "occupation_group_id": uri_key,
                "occupation_group_label": profile["occupation_group_label"],
                "profile": profile["profile"]
            }

        # fallback: try raw code if ever needed
        profile = self.esco_matrix_profiles.get((sheet_name, group_code))
        if profile:
            return {
                "sheet_name": sheet_name,
                "occupation_group_id": group_code,
                "occupation_group_label": profile["occupation_group_label"],
                "profile": profile["profile"]
            }

        return None

    def build_and_summarize_official_matrix_sector_skillgroups(
            self,
            jobs: List[dict],
            sector_level: str = "isco_group",
            skill_group_level: int = 1,
            occupation_level: int = 1,
            top_k: int = 20,
            reset: bool = True
    ):
        self.build_official_matrix_sector_skillgroup_profile(
            jobs=jobs,
            sector_level=sector_level,
            skill_group_level=skill_group_level,
            occupation_level=occupation_level,
            reset=reset
        )
        return self.summarize_official_matrix_sector_skillgroups(top_k=top_k)
    def summarize_official_matrix_single_sector(
            self,
            sector_name: str,
            top_k: int = 20
    ):
        """
        Return a readable summary for one sector using the official ESCO matrix profile.
        """
        return self.get_official_matrix_groups_for_sector(
            sector_name=sector_name,
            top_k=top_k
        )
    def get_official_matrix_groups_for_sector(self, sector_name: str, top_k: int = 20):
        """
        Return the official ESCO matrix skill-group profile for a sector.
        """
        sector_name = str(sector_name).strip()
        counter = self.matrix_profiles.get(sector_name, Counter())
        total_group_mentions = sum(counter.values())

        top_groups = []
        for group_id, count in counter.most_common(top_k):
            top_groups.append({
                "group_id": group_id,
                "group_label": self.get_skill_group_label(group_id),
                "count": count,
                "frequency": round(count / total_group_mentions, 6) if total_group_mentions > 0 else 0.0
            })

        return {
            "sector": sector_name,
            "total_group_mentions": total_group_mentions,
            "unique_groups": len(counter),
            "top_groups": top_groups
        }
    def build_official_matrix_sector_skillgroup_profile(
        self,
        jobs: List[dict],
        sector_level: str = "isco_group",
        skill_group_level: int = 1,
        occupation_level: int = 1,
        reset: bool = True
    ):
        """
        Build a sector -> official ESCO skill-group profile using the workbook.

        For each job:
        - resolve occupation
        - resolve sector
        - get official matrix profile for that occupation group
        - accumulate profile values into the sector
        """
        if reset:
            self.matrix_profiles = defaultdict(Counter)

        for job in jobs:
            occ_ids = self.get_occupation_ids(job)
            if not occ_ids:
                continue

            for occ_id in occ_ids:
                group_code = self.get_occupation_group_id_for_matrix(
                    occ_id=occ_id,
                    occupation_level=occupation_level
                )

                if not group_code:
                    continue

                sector_name = self.get_sector_from_occupation(occ_id, level=sector_level)

                official = self.get_official_esco_profile_for_occupation(
                    occ_id=occ_id,
                    skill_group_level=skill_group_level,
                    occupation_level=occupation_level
                )
                if not official:
                    continue

                for skill_group_id, share in official["profile"].items():
                    self.matrix_profiles[sector_name][skill_group_id] += float(share)

        return self.matrix_profiles
    def summarize_official_matrix_sector_skillgroups(self, top_k: int = 20):
        """
        Build a readable summary of official ESCO matrix skill-group profiles per sector.
        """
        results = []

        for sector_name, counter in self.matrix_profiles.items():
            base = self._read_group_counter(counter, top_k=top_k)
            results.append({
                "sector": sector_name,
                "total_group_mentions": base["total_mentions"],
                "unique_groups": base["unique_groups"],
                "top_groups": base["top_groups"]
            })

        return sorted(results, key=lambda x: x["total_group_mentions"], reverse=True)

    def compare_all_group_profiles_for_sector(
        self,
        sector_name: str,
        top_k: int = 20
    ):
        """
        Compare observed vs canonical vs official matrix group profiles for one sector.
        """
        sector_name = str(sector_name).strip()

        observed_counter = self.sector_skillgroup_observed.get(sector_name, Counter())
        canonical_counter = self.sector_skillgroup_canonical.get(sector_name, Counter())
        official_counter = self.matrix_profiles.get(sector_name, Counter())

        observed = self._read_group_counter(observed_counter, top_k=top_k)
        canonical = self._read_group_counter(canonical_counter, top_k=top_k)
        official = self._read_group_counter(official_counter, top_k=top_k)

        return {
            "sector": sector_name,
            "observed_groups": {
                "total_group_mentions": observed["total_mentions"],
                "unique_groups": observed["unique_groups"],
                "top_groups": observed["top_groups"]
            },
            "canonical_groups": {
                "total_group_mentions": canonical["total_mentions"],
                "unique_groups": canonical["unique_groups"],
                "top_groups": canonical["top_groups"]
            },
            "official_matrix_groups": {
                "total_group_mentions": official["total_mentions"],
                "unique_groups": official["unique_groups"],
                "top_groups": official["top_groups"]
            }
        }

engine = ProjectorEngine()
engine.load_local_esco_support()
engine.load_official_esco_matrix()

@app.post("/projector/analyze-skills", response_model=ProjectorResponse)
async def analyze_skills(
        keywords: Optional[List[str]] = Form(None),
        locations: Optional[List[str]] = Form(None),
        min_date: str = Form(...),
        max_date: str = Form(...),
        page: int = Form(1),
        page_size: int = Form(50),
        demo: bool = Form(False),
        include_sectoral: bool = Form(False),
        skill_group_level: int = Form(1),
        occupation_level: int = Form(1),
):
    """
       Executes a full labor market analysis based on user-defined filters.

       This endpoint orchestrates the entire pipeline:
           1. Fetch job postings from Tracker API
           2. Enrich skills and sectors
           3. Compute aggregated statistics
           4. Generate structured insights

       Args:
           keywords (List[str], optional): Search keywords for job filtering.
           min_date (str, optional): Start date (YYYY-MM-DD).
           max_date (str, optional): End date (YYYY-MM-DD).
           location_code (str, optional): Geographic filter (ISO/NUTS).
           occupation_ids (List[str], optional): Sector filter (ESCO).

       Returns:
           ProjectorResponse:
               - status
               - dimension_summary
               - insights (skills, employers, job titles, etc.)

       Notes:
           - Supports large-scale analysis (tens of thousands of jobs)
           - Uses caching for performance
           - Can be interrupted via `/projector/stop`
       """
    engine.stop_requested = False

    # Costruzione payload pulita
    payload = {
        "keywords": keywords,
        "location_code": locations,
        "min_upload_date": min_date,
        "max_upload_date": max_date
    }
    clean_payload = {k: v for k, v in payload.items() if v is not None}

    # FETCH UNICO
    raw = await engine.fetch_all_jobs(clean_payload)

    sectoral_data = None
    if include_sectoral:
        sectoral_data = engine.build_sectoral_intelligence(
            jobs=raw,
            sector_level="isco_group",
            skill_group_level=skill_group_level,
            occupation_level=occupation_level,
            resolve_labels=True,
            top_k_skills=10,
            top_k_groups=10,
            reset=True
        )

    # FIX: Se non ci sono job, restituiamo subito la struttura coerente
    if not raw:
        return {
            "status": "completed",
            "dimension_summary": {"jobs_analyzed": 0, "geo_breakdown": []},
            "insights": engine._empty_insights_p1()  # <--- Coerenza qui
        }

    # Traduzione settori (Phase 1)
    all_occs = []
    all_skills = []
    for j in raw:
        all_occs.extend(engine.get_occupation_ids(j))
        all_skills.extend(j.get("skills", []))  # <--- Aggiungiamo questo!

    # Add canonical ESCO skills from occupation-skill relations so their labels are resolved too
    canonical_skill_ids = set()
    for occ_id in set(all_occs):
        canonical_skill_ids.update(engine.occ_skill_relations.get(occ_id, set()))

    all_skills.extend(list(canonical_skill_ids))

    occ_uris = list(set(all_occs))  # Rimuove i duplicati

    await engine.fetch_occupation_labels(occ_uris)
    await engine.fetch_skill_names(list(set(all_skills)))  # <--- Traduciamo tutto il set

    # ------------------------------------------------------------------
    # Resolve ALL skills needed by the sectoral layer before building it:
    # 1. observed skills from raw jobs
    # 2. canonical ESCO skills linked to the occupations found in raw jobs
    # ------------------------------------------------------------------
    observed_skill_ids = set()
    canonical_skill_ids = set()

    for j in raw:
        for s in j.get("skills", []):
            s = str(s).strip()
            if s:
                observed_skill_ids.add(s)

    for occ_id in occ_uris:
        occ_id = str(occ_id).strip()
        if not occ_id:
            continue
        canonical_skill_ids.update(engine.occ_skill_relations.get(occ_id, set()))

    all_skill_ids_to_resolve = list(observed_skill_ids | canonical_skill_ids)

    await engine.fetch_skill_names(all_skill_ids_to_resolve)
    # Analisi globale
    analysis = await engine.analyze_market_data(raw)

    # Trend in memoria (Single Fetch optimization)
    trends = await engine.calculate_trends_from_data(raw, min_date, max_date)

    regional_projections = engine.get_regional_projections(raw, demo=demo)

    sectoral_data = None
    if include_sectoral:
        sectoral_data = engine.build_sectoral_intelligence(
            jobs=raw,
            sector_level="isco_group",
            skill_group_level=skill_group_level,
            occupation_level=occupation_level,
            resolve_labels=True,
            top_k_skills=10,
            top_k_groups=10,
            reset=True
        )

    start = (page - 1) * page_size

    return {
        "status": "completed" if not engine.stop_requested else "stopped",
        "dimension_summary": {
            "jobs_analyzed": analysis["total_jobs"],
            "geo_breakdown": analysis["geo"]
        },
        "insights": {
            "ranking": analysis["rankings"]["skills"][start: start + page_size],
            "sectors": analysis["rankings"]["sectors"],
            "job_titles": analysis["rankings"]["job_titles"],
            "employers": analysis["rankings"]["employers"],
            "trends": trends,
            "regional": regional_projections,
            "sectoral": sectoral_data
        }
    }


@app.post("/projector/emerging-skills", response_model=EmergingSkillsResponse)
async def emerging_skills(min_date: str = Form(...), max_date: str = Form(...),
                          keywords: Optional[List[str]] = Form(None)):
    """
       Computes emerging and declining skill trends over a time period.

       The analysis splits the time window into two segments:
           - Period A (past)
           - Period B (recent)

       It then computes growth rates to identify:
           - Emerging skills (increasing demand)
           - Declining skills (decreasing demand)
           - New entries (not present in previous period)

       Args:
           min_date (str): Start date (YYYY-MM-DD).
           max_date (str): End date (YYYY-MM-DD).

       Returns:
           EmergingSkillsResponse:
               - market_health (global trend)
               - trends (per-skill analysis)

       Key Metric:
           Growth % = (B - A) / A * 100
       """
    engine.stop_requested = False
    res = await engine.calculate_smart_trends({"keywords": keywords} if keywords else {}, min_date, max_date)
    return {"status": "completed" if not engine.stop_requested else "stopped", "insights": res}


@app.post("/projector/stop", response_model=StopResponse)
async def stop():
    """
        Sends a stop signal to interrupt ongoing analysis tasks.

        This endpoint triggers a cooperative stop mechanism in the engine.
        The running process will terminate at the next safe checkpoint.

        Returns:
            dict:
                {"status": "stopping"}

        Notes:
            - Does not immediately kill execution
            - Safe for long-running operations
        """

    engine.request_stop()
    return {"status": "signal_sent"}


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000)
